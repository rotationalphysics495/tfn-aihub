"""
Schedule Parser Service

Business logic for CSV/Excel parsing, column header validation, per-row validation,
fuzzy asset matching, product matching, and preview assembly.

Story: 12.3 - Schedule Upload API
AC: #1 - CSV Upload Preview
AC: #3 - Excel Support
AC: #4 - Fuzzy Asset Matching
AC: #5 - Validation Errors
"""

import csv
import io
from datetime import date, datetime, timedelta
from difflib import get_close_matches
from typing import Dict, List, Optional, Tuple

from app.schemas.schedule import (
    RowValidationError,
    SchedulePreviewResponse,
    ScheduleRowPreview,
)

# Required column headers (normalized to lowercase with underscores)
REQUIRED_COLUMNS = {"date", "shift", "asset_name", "product_name", "scheduled_quantity"}

# Column name aliases mapping common variants to canonical names
COLUMN_ALIASES = {
    "quantity": "scheduled_quantity",
    "qty": "scheduled_quantity",
    "scheduled_qty": "scheduled_quantity",
    "asset": "asset_name",
    "product": "product_name",
}


def detect_file_format(filename: str, content_type: Optional[str] = None) -> Optional[str]:
    """Detect file format from filename extension (case-insensitive).

    Returns "csv", "xlsx", or None for unsupported types.
    """
    if filename:
        lower = filename.lower()
        if lower.endswith(".csv"):
            return "csv"
        if lower.endswith(".xlsx"):
            return "xlsx"
    return None


def normalize_headers(headers: List[str]) -> Dict[str, str]:
    """Normalize column headers to lowercase with underscores.

    Applies alias mapping so common variants (e.g., "quantity") map
    to canonical names (e.g., "scheduled_quantity").

    Returns a mapping of normalized_name -> original_name.
    """
    mapping = {}
    for h in headers:
        normalized = h.strip().lower().replace(" ", "_")
        # Apply alias if available
        normalized = COLUMN_ALIASES.get(normalized, normalized)
        mapping[normalized] = h
    return mapping


def validate_headers(headers: List[str]) -> List[str]:
    """Validate that all required column headers are present.

    Applies the same alias normalization as normalize_headers so that
    common variants like "quantity" satisfy "scheduled_quantity".

    Args:
        headers: List of header strings (will be normalized for comparison).

    Returns:
        List of missing required column names (empty if all present).
    """
    normalized = set()
    for h in headers:
        norm = h.strip().lower().replace(" ", "_")
        norm = COLUMN_ALIASES.get(norm, norm)
        normalized.add(norm)
    missing = []
    for req in sorted(REQUIRED_COLUMNS):
        if req not in normalized:
            missing.append(req)
    return missing


def parse_csv(file_bytes: bytes) -> List[Dict[str, str]]:
    """Parse CSV file bytes into a list of row dictionaries.

    Handles UTF-8 BOM encoding. Normalizes column headers to lowercase.

    Args:
        file_bytes: Raw bytes of the CSV file.

    Returns:
        List of dictionaries, one per data row, with normalized keys.
    """
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    if reader.fieldnames is None:
        return []

    # Build normalized header mapping
    header_map = normalize_headers(list(reader.fieldnames))

    rows = []
    for raw_row in reader:
        normalized_row = {}
        for norm_key, orig_key in header_map.items():
            value = raw_row.get(orig_key, "")
            normalized_row[norm_key] = value.strip() if value else ""
        rows.append(normalized_row)

    return rows


def parse_excel(file_bytes: bytes) -> List[Dict[str, str]]:
    """Parse Excel (.xlsx) file bytes into a list of row dictionaries.

    Reads only the first (active) sheet. Treats row 1 as headers.
    Uses data_only=True to get calculated values instead of formulas.

    Args:
        file_bytes: Raw bytes of the .xlsx file.

    Returns:
        List of dictionaries, one per data row, with normalized keys.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows()

    # First row is headers
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return []

    raw_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_row]
    header_map = normalize_headers(raw_headers)
    norm_keys = list(header_map.keys())

    rows = []
    for row in rows_iter:
        values = [cell.value for cell in row]
        # Skip completely empty rows
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in values):
            continue

        row_dict = {}
        for i, norm_key in enumerate(norm_keys):
            if i < len(values):
                val = values[i]
                row_dict[norm_key] = str(val).strip() if val is not None else ""
            else:
                row_dict[norm_key] = ""
        rows.append(row_dict)

    wb.close()
    return rows


def _parse_date(date_str: str) -> Optional[date]:
    """Try to parse a date string in ISO (YYYY-MM-DD) or US (MM/DD/YYYY) format."""
    date_str = date_str.strip()

    # ISO format: YYYY-MM-DD
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        pass

    # US format: MM/DD/YYYY
    try:
        return datetime.strptime(date_str, "%m/%d/%Y").date()
    except ValueError:
        pass

    return None


def validate_row(row: Dict[str, str], row_number: int) -> List[RowValidationError]:
    """Validate a single parsed row and return all validation errors.

    Validates:
    - date: valid format, within range (-90 to +365 days)
    - shift: non-empty
    - asset_name: non-empty
    - product_name: non-empty
    - scheduled_quantity: positive integer

    Args:
        row: Dictionary of field values (all as strings).
        row_number: 1-based row number for error reporting.

    Returns:
        List of RowValidationError objects (empty if valid).
    """
    errors = []
    today = date.today()

    # Validate date
    date_str = row.get("date", "").strip()
    if date_str:
        parsed = _parse_date(date_str)
        if parsed is None:
            errors.append(RowValidationError(
                row_number=row_number,
                field="date",
                message="Invalid date format. Use YYYY-MM-DD or MM/DD/YYYY.",
            ))
        else:
            min_date = today - timedelta(days=90)
            max_date = today + timedelta(days=365)
            if parsed < min_date:
                errors.append(RowValidationError(
                    row_number=row_number,
                    field="date",
                    message="Date is too far in the past (more than 90 days ago).",
                ))
            elif parsed > max_date:
                errors.append(RowValidationError(
                    row_number=row_number,
                    field="date",
                    message="Date is too far in the future (more than 365 days ahead).",
                ))
    else:
        errors.append(RowValidationError(
            row_number=row_number,
            field="date",
            message="Date is required.",
        ))

    # Validate shift
    shift = row.get("shift", "").strip()
    if not shift:
        errors.append(RowValidationError(
            row_number=row_number,
            field="shift",
            message="Shift must be non-empty.",
        ))

    # Validate asset_name
    asset_name = row.get("asset_name", "").strip()
    if not asset_name:
        errors.append(RowValidationError(
            row_number=row_number,
            field="asset_name",
            message="Asset name must be non-empty.",
        ))

    # Validate product_name
    product_name = row.get("product_name", "").strip()
    if not product_name:
        errors.append(RowValidationError(
            row_number=row_number,
            field="product_name",
            message="Product name must be non-empty.",
        ))

    # Validate scheduled_quantity
    qty_str = row.get("scheduled_quantity", "").strip()
    if qty_str:
        try:
            qty_float = float(qty_str)
            if qty_float != int(qty_float):
                errors.append(RowValidationError(
                    row_number=row_number,
                    field="scheduled_quantity",
                    message="Scheduled quantity must be a positive integer.",
                ))
            else:
                qty_int = int(qty_float)
                if qty_int <= 0:
                    errors.append(RowValidationError(
                        row_number=row_number,
                        field="scheduled_quantity",
                        message="Scheduled quantity must be a positive integer greater than 0.",
                    ))
        except ValueError:
            errors.append(RowValidationError(
                row_number=row_number,
                field="scheduled_quantity",
                message="Scheduled quantity must be a positive integer.",
            ))
    else:
        errors.append(RowValidationError(
            row_number=row_number,
            field="scheduled_quantity",
            message="Scheduled quantity is required.",
        ))

    return errors


def match_asset(
    input_name: str,
    asset_names_map: Dict[str, str],
    cutoff: float = 0.6,
) -> Dict:
    """Match an input asset name against known assets.

    Logic:
    1. Exact case-insensitive match -> "matched" with asset_id
    2. Fuzzy close matches -> "suggested" with suggestions list
    3. No matches -> "unmatched"

    Args:
        input_name: The asset name from the uploaded file.
        asset_names_map: Dict mapping asset_name -> asset_id (UUID string).
        cutoff: Minimum similarity ratio for fuzzy matching (0.0 to 1.0).

    Returns:
        Dict with keys: status, asset_id, suggestions
    """
    # 1. Exact case-insensitive match
    input_lower = input_name.strip().lower()
    for name, asset_id in asset_names_map.items():
        if name.lower() == input_lower:
            return {"status": "matched", "asset_id": asset_id, "suggestions": []}

    # 2. Fuzzy matching
    asset_names = list(asset_names_map.keys())
    close_matches = get_close_matches(input_name, asset_names, n=3, cutoff=cutoff)

    if close_matches:
        return {"status": "suggested", "asset_id": None, "suggestions": close_matches}

    # 3. No match
    return {"status": "unmatched", "asset_id": None, "suggestions": []}


def match_products(
    product_names: List[str],
    existing_products: Dict[str, str],
) -> Dict[str, Optional[str]]:
    """Match product names against existing products.

    Args:
        product_names: List of product names from the uploaded file.
        existing_products: Dict mapping product_name -> product_id.

    Returns:
        Dict mapping product_name -> product_id (or None if new).
    """
    result = {}
    for name in product_names:
        result[name] = existing_products.get(name)
    return result


def assemble_preview(
    parsed_rows: List[Dict[str, str]],
    asset_names_map: Dict[str, str],
    existing_products: Dict[str, str],
) -> SchedulePreviewResponse:
    """Assemble a preview response from parsed rows, asset matches, and product matches.

    Args:
        parsed_rows: List of normalized row dicts from CSV/Excel parsing.
        asset_names_map: Dict mapping asset_name -> asset_id.
        existing_products: Dict mapping product_name -> product_id.

    Returns:
        SchedulePreviewResponse with all rows, match statuses, and error flags.
    """
    preview_rows = []
    matched_assets_set = set()
    matched_products_set = set()
    new_products_set = set()
    has_errors = False

    for i, row in enumerate(parsed_rows):
        row_number = i + 1
        errors = validate_row(row, row_number)

        # Asset matching
        asset_name = row.get("asset_name", "").strip()
        asset_result = match_asset(asset_name, asset_names_map)

        if asset_result["status"] == "matched":
            matched_assets_set.add(asset_name)
        elif asset_result["status"] == "unmatched":
            errors.append(RowValidationError(
                row_number=row_number,
                field="asset_name",
                message=f"Asset '{asset_name}' not found and no close matches available.",
            ))

        # Product matching
        product_name = row.get("product_name", "").strip()
        product_id = existing_products.get(product_name)
        is_new_product = product_id is None and product_name != ""

        if product_id:
            matched_products_set.add(product_name)
        elif is_new_product:
            new_products_set.add(product_name)

        if errors:
            has_errors = True

        preview_rows.append(ScheduleRowPreview(
            row_number=row_number,
            date=row.get("date", ""),
            shift=row.get("shift", ""),
            asset_name=asset_name,
            product_name=product_name,
            scheduled_quantity=row.get("scheduled_quantity", ""),
            asset_match_status=asset_result["status"],
            asset_id=asset_result["asset_id"],
            suggestions=asset_result["suggestions"],
            product_id=product_id,
            is_new_product=is_new_product,
            errors=errors,
        ))

    # If any row has suggested status (not matched), flag as having errors
    if any(r.asset_match_status == "suggested" for r in preview_rows):
        has_errors = True

    return SchedulePreviewResponse(
        parsed_rows_count=len(parsed_rows),
        rows=preview_rows,
        has_errors=has_errors,
        matched_assets=sorted(matched_assets_set),
        matched_products=sorted(matched_products_set),
        new_products=sorted(new_products_set),
    )
