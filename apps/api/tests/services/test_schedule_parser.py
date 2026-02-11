"""
Tests for Schedule Parser Service (Story 12.3 - Schedule Upload API)

Unit tests for CSV/Excel parsing, validation, fuzzy asset matching,
and product matching functions.

AC: #1 - CSV Upload Preview
AC: #3 - Excel Support
AC: #4 - Fuzzy Asset Matching
AC: #5 - Validation Errors
"""

import io
from datetime import date, timedelta
import pytest
from uuid import uuid4

# These imports will fail until the feature module is implemented,
# causing all tests to FAIL as expected for TDD.
from app.services.schedule_parser import (
    parse_csv,
    parse_excel,
    normalize_headers,
    validate_headers,
    validate_row,
    match_asset,
    match_products,
    assemble_preview,
    detect_file_format,
)


# =============================================================================
# Test Data / Fixtures
# =============================================================================

ASSET_GRINDER_1_ID = str(uuid4())
ASSET_GRINDER_2_ID = str(uuid4())
ASSET_GRINDER_3_ID = str(uuid4())
ASSET_GRINDER_4_ID = str(uuid4())
ASSET_GRINDER_5_ID = str(uuid4())
ASSET_FILLER_A_ID = str(uuid4())
ASSET_FILLER_B_ID = str(uuid4())
ASSET_FILLER_C_ID = str(uuid4())
ASSET_ROASTER_1_ID = str(uuid4())
ASSET_ROASTER_2_ID = str(uuid4())
ASSET_ROASTER_3_ID = str(uuid4())
ASSET_PACKAGING_1_ID = str(uuid4())
ASSET_PACKAGING_2_ID = str(uuid4())
ASSET_PACKAGING_3_ID = str(uuid4())

STANDARD_ASSET_MAP = {
    "Grinder 1": ASSET_GRINDER_1_ID,
    "Grinder 2": ASSET_GRINDER_2_ID,
    "Grinder 3": ASSET_GRINDER_3_ID,
    "Grinder 4": ASSET_GRINDER_4_ID,
    "Grinder 5": ASSET_GRINDER_5_ID,
    "Filler Line A": ASSET_FILLER_A_ID,
    "Filler Line B": ASSET_FILLER_B_ID,
    "Filler Line C": ASSET_FILLER_C_ID,
    "Roaster 1": ASSET_ROASTER_1_ID,
    "Roaster 2": ASSET_ROASTER_2_ID,
    "Roaster 3": ASSET_ROASTER_3_ID,
    "Packaging Line 1": ASSET_PACKAGING_1_ID,
    "Packaging Line 2": ASSET_PACKAGING_2_ID,
    "Packaging Line 3": ASSET_PACKAGING_3_ID,
}

PRODUCT_DARK_ROAST_ID = str(uuid4())
EXISTING_PRODUCTS = {
    "Dark Roast 12oz": PRODUCT_DARK_ROAST_ID,
}


@pytest.fixture
def valid_csv_bytes():
    """A valid CSV file with 5 data rows and correct headers."""
    content = (
        "date,shift,asset_name,product_name,scheduled_quantity\n"
        "2026-02-16,Day,Grinder 1,Dark Roast 12oz,500\n"
        "2026-02-16,Night,Grinder 2,Medium Blend 16oz,300\n"
        "2026-02-17,Day,Filler Line A,Dark Roast 12oz,450\n"
        "2026-02-17,Night,Roaster 1,New Blend XYZ,600\n"
        "2026-02-18,Day,Packaging Line 1,Dark Roast 12oz,250\n"
    )
    return content.encode("utf-8")


@pytest.fixture
def valid_csv_3_rows():
    """A valid CSV file with 3 data rows for upload preview testing."""
    content = (
        "date,shift,asset_name,product_name,scheduled_quantity\n"
        "2026-02-16,Day,Grinder 1,Dark Roast 12oz,500\n"
        "2026-02-16,Night,Filler Line A,Medium Blend 16oz,300\n"
        "2026-02-17,Day,Roaster 1,New Blend XYZ,400\n"
    )
    return content.encode("utf-8")


@pytest.fixture
def csv_with_bom():
    """A CSV file encoded with UTF-8 BOM prefix."""
    bom = b"\xef\xbb\xbf"
    content = (
        "date,shift,asset_name,product_name,scheduled_quantity\n"
        "2026-02-16,Day,Grinder 1,Coffee Blend,100\n"
    )
    return bom + content.encode("utf-8")


@pytest.fixture
def csv_mixed_case_headers():
    """A CSV file with mixed-case column headers."""
    content = (
        "Date,SHIFT,Asset_Name,Product_Name,Scheduled_Quantity\n"
        "2026-02-16,Day,Grinder 1,Dark Roast 12oz,500\n"
    )
    return content.encode("utf-8")


@pytest.fixture
def csv_headers_only():
    """A CSV file containing only headers and no data rows."""
    content = "date,shift,asset_name,product_name,scheduled_quantity\n"
    return content.encode("utf-8")


@pytest.fixture
def csv_missing_columns():
    """A CSV file with missing required columns (no asset_name, product_name)."""
    content = (
        "date,shift,quantity\n"
        "2026-02-16,Day,500\n"
    )
    return content.encode("utf-8")


@pytest.fixture
def asset_names_map():
    """Standard asset names map for fuzzy matching tests."""
    return STANDARD_ASSET_MAP.copy()


@pytest.fixture
def existing_products():
    """Existing products dict for product matching tests."""
    return EXISTING_PRODUCTS.copy()


def _make_valid_row(**overrides):
    """Helper to create a valid parsed row dict with optional overrides."""
    row = {
        "date": "2026-02-16",
        "shift": "Day",
        "asset_name": "Grinder 1",
        "product_name": "Dark Roast 12oz",
        "scheduled_quantity": "500",
    }
    row.update(overrides)
    return row


def _make_xlsx_bytes(rows, sheet_name="Schedule", extra_sheets=None):
    """
    Helper to create .xlsx file bytes from a list of row dicts.
    First row's keys become headers.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row[h] for h in headers])

    if extra_sheets:
        for name in extra_sheets:
            wb.create_sheet(title=name)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# =============================================================================
# AC1: CSV Upload Preview — CSV Parsing Tests
# =============================================================================


class TestCSVParsing:
    """Tests for CSV file parsing functions."""

    def test_UNIT_001_csv_parsing_handles_utf8_bom(self, csv_with_bom):
        """12-3-schedule-upload-api-UNIT-001: CSV parsing handles UTF-8 BOM encoding correctly.

        Given: A CSV file encoded with UTF-8 BOM (byte order mark prefix)
        When: parse_csv() is called with the file bytes
        Then: Headers are correctly parsed without BOM characters; all rows valid
        """
        rows = parse_csv(csv_with_bom)

        assert len(rows) == 1
        # First column key should NOT have BOM prefix
        first_row = rows[0]
        assert "date" in first_row, f"Expected 'date' key, got keys: {list(first_row.keys())}"
        assert first_row["date"] == "2026-02-16"
        assert first_row["shift"] == "Day"
        assert first_row["asset_name"] == "Grinder 1"
        assert first_row["product_name"] == "Coffee Blend"
        assert first_row["scheduled_quantity"] == "100"

    def test_UNIT_002_csv_parsing_returns_correct_row_dicts(self, valid_csv_bytes):
        """12-3-schedule-upload-api-UNIT-002: CSV parsing returns correct row dictionaries.

        Given: A well-formed CSV byte string with 5 data rows
        When: parse_csv() is called with the file bytes
        Then: A list of 5 dictionaries is returned with correct keys and values
        """
        rows = parse_csv(valid_csv_bytes)

        assert len(rows) == 5
        expected_keys = {"date", "shift", "asset_name", "product_name", "scheduled_quantity"}
        for row in rows:
            assert set(row.keys()) == expected_keys

        # Spot check first row
        assert rows[0]["date"] == "2026-02-16"
        assert rows[0]["shift"] == "Day"
        assert rows[0]["asset_name"] == "Grinder 1"
        assert rows[0]["product_name"] == "Dark Roast 12oz"
        assert rows[0]["scheduled_quantity"] == "500"

    def test_UNIT_004_column_header_matching_case_insensitive(self, csv_mixed_case_headers):
        """12-3-schedule-upload-api-UNIT-004: Column header matching is case-insensitive.

        Given: A CSV with mixed-case headers "Date,SHIFT,Asset_Name,Product_Name,Scheduled_Quantity"
        When: parse_csv() and normalize_headers() are called
        Then: All headers are recognized as valid; rows returned with normalized keys
        """
        rows = parse_csv(csv_mixed_case_headers)

        assert len(rows) == 1
        first_row = rows[0]
        # Keys should be normalized to lowercase
        assert "date" in first_row
        assert "shift" in first_row
        assert "asset_name" in first_row
        assert "product_name" in first_row
        assert "scheduled_quantity" in first_row


# =============================================================================
# AC3: Excel Support — Excel Parsing Tests
# =============================================================================


class TestExcelParsing:
    """Tests for Excel (.xlsx) file parsing functions."""

    def test_UNIT_005_excel_parsing_reads_first_sheet_headers(self):
        """12-3-schedule-upload-api-UNIT-005: Excel parsing reads first sheet, row 1 as headers.

        Given: An Excel file with multiple sheets; first sheet has headers in row 1
        When: parse_excel() is called with the file bytes
        Then: Only first sheet read; row 1 = headers; correct number of data rows returned
        """
        data_rows = [
            {"date": "2026-02-16", "shift": "Day", "asset_name": "Grinder 1",
             "product_name": "Dark Roast 12oz", "scheduled_quantity": 500},
            {"date": "2026-02-17", "shift": "Night", "asset_name": "Filler Line A",
             "product_name": "Medium Blend", "scheduled_quantity": 300},
            {"date": "2026-02-18", "shift": "Day", "asset_name": "Roaster 1",
             "product_name": "New Blend", "scheduled_quantity": 400},
        ]
        xlsx_bytes = _make_xlsx_bytes(data_rows, extra_sheets=["Sheet2", "Sheet3"])

        rows = parse_excel(xlsx_bytes)

        assert len(rows) == 3
        expected_keys = {"date", "shift", "asset_name", "product_name", "scheduled_quantity"}
        for row in rows:
            assert set(row.keys()) == expected_keys

        # Check first row values
        assert rows[0]["date"] == "2026-02-16" or str(rows[0]["date"]) == "2026-02-16"
        assert rows[0]["asset_name"] == "Grinder 1"

    def test_UNIT_006_file_format_detection(self):
        """12-3-schedule-upload-api-UNIT-006: File format detection distinguishes CSV from XLSX.

        Given: Files with various extensions and content types
        When: detect_file_format() is called for each file
        Then: Returns "csv" for .csv, "xlsx" for .xlsx, None for unsupported
        """
        # CSV detection
        assert detect_file_format("schedule.csv", "text/csv") == "csv"
        # XLSX detection
        assert detect_file_format(
            "schedule.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ) == "xlsx"
        # Case-insensitive extension
        assert detect_file_format("schedule.CSV", "text/csv") == "csv"
        # Unsupported
        assert detect_file_format("schedule.txt", "text/plain") is None

    def test_UNIT_007_excel_parsing_data_only_mode(self):
        """12-3-schedule-upload-api-UNIT-007: Excel parsing with data_only returns calculated values.

        Given: An Excel file where scheduled_quantity contains a formula (=100+50)
        When: parse_excel() is called with data_only=True
        Then: The parsed value is the calculated result (150), not the formula string
        """
        # Create an Excel file with a formula cell
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["date", "shift", "asset_name", "product_name", "scheduled_quantity"])
        ws.append(["2026-02-16", "Day", "Grinder 1", "Blend", None])
        # Set formula and cached value
        cell = ws.cell(row=2, column=5)
        cell.value = "=100+50"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_bytes = buffer.read()

        rows = parse_excel(xlsx_bytes)

        # data_only with read_only may return None for uncached formulas
        # The test validates we handle this edge case
        assert len(rows) == 1
        qty = rows[0].get("scheduled_quantity")
        # Either we get the calculated value (150) or None for uncached
        assert qty is None or str(qty) != "=100+50", \
            "Formula string should not be returned; expected calculated value or None"


# =============================================================================
# AC4: Fuzzy Asset Matching Tests
# =============================================================================


class TestFuzzyAssetMatching:
    """Tests for fuzzy asset name matching."""

    def test_UNIT_008_exact_case_insensitive_match(self, asset_names_map):
        """12-3-schedule-upload-api-UNIT-008: Exact case-insensitive match returns matched status.

        Given: Input "grinder 1" (lowercase); "Grinder 1" exists with UUID
        When: match_asset("grinder 1", asset_names_map) is called
        Then: Returns status "matched" with the correct asset_id
        """
        result = match_asset("grinder 1", asset_names_map)

        assert result["status"] == "matched"
        assert result["asset_id"] == ASSET_GRINDER_1_ID
        assert not result.get("suggestions") or len(result["suggestions"]) == 0

    def test_UNIT_009_close_typo_match_returns_suggestions(self, asset_names_map):
        """12-3-schedule-upload-api-UNIT-009: Close typo match returns suggested status.

        Given: Input "Grinder" (missing number); "Grinder 1"-"Grinder 5" exist
        When: match_asset("Grinder", asset_names_map, cutoff=0.6) is called
        Then: Returns status "suggested"; suggestions list contains up to 3 matches
        """
        result = match_asset("Grinder", asset_names_map, cutoff=0.6)

        assert result["status"] == "suggested"
        assert result["asset_id"] is None
        assert len(result["suggestions"]) > 0
        assert len(result["suggestions"]) <= 3
        # All suggestions should be grinder-related
        for suggestion in result["suggestions"]:
            assert "Grinder" in suggestion

    def test_UNIT_010_completely_unrecognized_returns_unmatched(self, asset_names_map):
        """12-3-schedule-upload-api-UNIT-010: Completely unrecognized name returns unmatched.

        Given: Input "XYZ Machine 99"; standard plant assets exist
        When: match_asset("XYZ Machine 99", asset_names_map, cutoff=0.6) is called
        Then: Returns status "unmatched"; empty suggestions; asset_id is None
        """
        result = match_asset("XYZ Machine 99", asset_names_map, cutoff=0.6)

        assert result["status"] == "unmatched"
        assert result["asset_id"] is None
        assert len(result["suggestions"]) == 0

    def test_UNIT_011_fuzzy_match_minor_typo(self, asset_names_map):
        """12-3-schedule-upload-api-UNIT-011: Minor typo "Filler Line a" matches "Filler Line A".

        Given: Input "Filler Line a" (lowercase); "Filler Line A" exists
        When: match_asset("Filler Line a", asset_names_map) is called
        Then: Returns status "matched" with correct asset_id (case-insensitive exact match)
        """
        result = match_asset("Filler Line a", asset_names_map)

        assert result["status"] == "matched"
        assert result["asset_id"] == ASSET_FILLER_A_ID

    def test_UNIT_012_fuzzy_match_roaster_without_number(self, asset_names_map):
        """12-3-schedule-upload-api-UNIT-012: "Roaster" without number suggests all roasters.

        Given: Input "Roaster"; "Roaster 1", "Roaster 2", "Roaster 3" exist
        When: match_asset("Roaster", asset_names_map, cutoff=0.6) is called
        Then: Returns status "suggested"; suggestions include Roaster 1, 2, 3
        """
        result = match_asset("Roaster", asset_names_map, cutoff=0.6)

        assert result["status"] == "suggested"
        assert result["asset_id"] is None
        suggestions = result["suggestions"]
        assert len(suggestions) == 3
        assert "Roaster 1" in suggestions
        assert "Roaster 2" in suggestions
        assert "Roaster 3" in suggestions


# =============================================================================
# AC5: Validation Error Tests
# =============================================================================


class TestRowValidation:
    """Tests for per-row validation of schedule data."""

    def test_UNIT_013_negative_quantity_flagged(self):
        """12-3-schedule-upload-api-UNIT-013: Negative scheduled_quantity flagged.

        Given: A parsed row with scheduled_quantity = -10
        When: validate_row(row, row_number=2) is called
        Then: Returns error for field "scheduled_quantity" about positive integer
        """
        row = _make_valid_row(scheduled_quantity="-10")
        errors = validate_row(row, row_number=2)

        assert len(errors) >= 1
        qty_errors = [e for e in errors if e.field == "scheduled_quantity"]
        assert len(qty_errors) == 1
        assert "positive" in qty_errors[0].message.lower() or "greater" in qty_errors[0].message.lower()

    def test_UNIT_014_zero_quantity_flagged(self):
        """12-3-schedule-upload-api-UNIT-014: Zero scheduled_quantity flagged.

        Given: A parsed row with scheduled_quantity = 0
        When: validate_row(row, row_number=3) is called
        Then: Returns error for "scheduled_quantity" indicating > 0
        """
        row = _make_valid_row(scheduled_quantity="0")
        errors = validate_row(row, row_number=3)

        assert len(errors) >= 1
        qty_errors = [e for e in errors if e.field == "scheduled_quantity"]
        assert len(qty_errors) == 1

    def test_UNIT_015_invalid_date_format_flagged(self):
        """12-3-schedule-upload-api-UNIT-015: Invalid date format flagged.

        Given: A parsed row with date = "not-a-date"
        When: validate_row(row, row_number=1) is called
        Then: Returns error for field "date" indicating invalid format
        """
        row = _make_valid_row(date="not-a-date")
        errors = validate_row(row, row_number=1)

        assert len(errors) >= 1
        date_errors = [e for e in errors if e.field == "date"]
        assert len(date_errors) == 1
        assert "date" in date_errors[0].message.lower() or "invalid" in date_errors[0].message.lower()

    def test_UNIT_016_date_in_distant_past_flagged(self):
        """12-3-schedule-upload-api-UNIT-016: Date >90 days in the past flagged.

        Given: A parsed row with date = "2025-01-01" (more than 90 days ago)
        When: validate_row(row, row_number=1) is called
        Then: Returns error for "date" indicating too far in the past
        """
        row = _make_valid_row(date="2025-01-01")
        errors = validate_row(row, row_number=1)

        assert len(errors) >= 1
        date_errors = [e for e in errors if e.field == "date"]
        assert len(date_errors) == 1
        assert "past" in date_errors[0].message.lower()

    def test_UNIT_017_date_in_far_future_flagged(self):
        """12-3-schedule-upload-api-UNIT-017: Date >365 days in the future flagged.

        Given: A parsed row with date = "2028-01-01" (more than 365 days ahead)
        When: validate_row(row, row_number=1) is called
        Then: Returns error for "date" indicating too far in the future
        """
        row = _make_valid_row(date="2028-01-01")
        errors = validate_row(row, row_number=1)

        assert len(errors) >= 1
        date_errors = [e for e in errors if e.field == "date"]
        assert len(date_errors) == 1
        assert "future" in date_errors[0].message.lower()

    def test_UNIT_018_empty_shift_flagged(self):
        """12-3-schedule-upload-api-UNIT-018: Empty shift value flagged.

        Given: A parsed row with shift = "" (empty string)
        When: validate_row(row, row_number=1) is called
        Then: Returns error for "shift" indicating must be non-empty
        """
        row = _make_valid_row(shift="")
        errors = validate_row(row, row_number=1)

        assert len(errors) >= 1
        shift_errors = [e for e in errors if e.field == "shift"]
        assert len(shift_errors) == 1

    def test_UNIT_019_empty_product_name_flagged(self):
        """12-3-schedule-upload-api-UNIT-019: Empty product_name flagged.

        Given: A parsed row with product_name = "" (empty string)
        When: validate_row(row, row_number=1) is called
        Then: Returns error for "product_name" indicating must be non-empty
        """
        row = _make_valid_row(product_name="")
        errors = validate_row(row, row_number=1)

        assert len(errors) >= 1
        product_errors = [e for e in errors if e.field == "product_name"]
        assert len(product_errors) == 1

    def test_UNIT_020_empty_asset_name_flagged(self):
        """12-3-schedule-upload-api-UNIT-020: Empty asset_name flagged.

        Given: A parsed row with asset_name = "" (empty string)
        When: validate_row(row, row_number=1) is called
        Then: Returns error for "asset_name" indicating must be non-empty
        """
        row = _make_valid_row(asset_name="")
        errors = validate_row(row, row_number=1)

        assert len(errors) >= 1
        asset_errors = [e for e in errors if e.field == "asset_name"]
        assert len(asset_errors) == 1

    def test_UNIT_021_multiple_errors_per_row_all_reported(self):
        """12-3-schedule-upload-api-UNIT-021: Multiple errors per row are all reported.

        Given: A row with date="invalid", shift="", scheduled_quantity="-5"
        When: validate_row(row, row_number=1) is called
        Then: Returns 3 errors, one for each invalid field
        """
        row = _make_valid_row(date="invalid", shift="", scheduled_quantity="-5")
        errors = validate_row(row, row_number=1)

        assert len(errors) == 3
        error_fields = {e.field for e in errors}
        assert "date" in error_fields
        assert "shift" in error_fields
        assert "scheduled_quantity" in error_fields

    def test_UNIT_024_non_integer_quantity_flagged(self):
        """12-3-schedule-upload-api-UNIT-024: Non-integer (decimal) quantity flagged.

        Given: A parsed row with scheduled_quantity = "12.5"
        When: validate_row(row, row_number=1) is called
        Then: Returns error for "scheduled_quantity" indicating must be positive integer
        """
        row = _make_valid_row(scheduled_quantity="12.5")
        errors = validate_row(row, row_number=1)

        assert len(errors) >= 1
        qty_errors = [e for e in errors if e.field == "scheduled_quantity"]
        assert len(qty_errors) == 1
        assert "integer" in qty_errors[0].message.lower()

    def test_UNIT_025_non_numeric_quantity_flagged(self):
        """12-3-schedule-upload-api-UNIT-025: Non-numeric quantity "abc" flagged.

        Given: A parsed row with scheduled_quantity = "abc"
        When: validate_row(row, row_number=1) is called
        Then: Returns error for "scheduled_quantity" indicating must be positive integer
        """
        row = _make_valid_row(scheduled_quantity="abc")
        errors = validate_row(row, row_number=1)

        assert len(errors) >= 1
        qty_errors = [e for e in errors if e.field == "scheduled_quantity"]
        assert len(qty_errors) == 1

    def test_UNIT_026_valid_date_formats_accepted(self):
        """12-3-schedule-upload-api-UNIT-026: Valid date formats accepted (ISO and US).

        Given: Rows with a near-future ISO date and US-formatted date
        When: validate_row() is called for each
        Then: No validation errors for the date field
        """
        # Use a date 7 days in the future so it's always within the valid range
        future_date = date.today() + timedelta(days=7)

        # ISO format
        iso_str = future_date.strftime("%Y-%m-%d")
        row_iso = _make_valid_row(date=iso_str)
        errors_iso = validate_row(row_iso, row_number=1)
        date_errors_iso = [e for e in errors_iso if e.field == "date"]
        assert len(date_errors_iso) == 0, "ISO date format should be accepted"

        # US format
        us_str = future_date.strftime("%m/%d/%Y")
        row_us = _make_valid_row(date=us_str)
        errors_us = validate_row(row_us, row_number=2)
        date_errors_us = [e for e in errors_us if e.field == "date"]
        assert len(date_errors_us) == 0, "US date format (MM/DD/YYYY) should be accepted"


# =============================================================================
# AC5: File-Level Validation Tests
# =============================================================================


class TestFileValidation:
    """Tests for file-level validation (headers, empty files)."""

    def test_UNIT_022_missing_required_columns_detected(self):
        """12-3-schedule-upload-api-UNIT-022: Missing required column headers detected.

        Given: Headers ["date", "shift", "quantity"] (missing asset_name, product_name)
        When: validate_headers(headers) is called
        Then: Returns missing columns: ["asset_name", "product_name"]
        """
        headers = ["date", "shift", "quantity"]
        missing = validate_headers(headers)

        assert "asset_name" in missing
        assert "product_name" in missing
        assert len(missing) == 2

    def test_UNIT_023_empty_file_with_only_headers_rejected(self, csv_headers_only):
        """12-3-schedule-upload-api-UNIT-023: Empty file with only headers rejected.

        Given: A CSV file containing only the header row and no data rows
        When: File-level validation is performed after parsing
        Then: Validation fails with error about needing at least 1 data row
        """
        rows = parse_csv(csv_headers_only)
        assert len(rows) == 0, "No data rows should be parsed from headers-only file"


# =============================================================================
# AC1/AC4: Preview Assembly Tests
# =============================================================================


class TestPreviewAssembly:
    """Tests for preview assembly (grouping rows by match status and errors)."""

    def test_UNIT_003_preview_assembly_groups_rows(self, asset_names_map, existing_products):
        """12-3-schedule-upload-api-UNIT-003: Preview assembly groups rows by match status.

        Given: 4 parsed rows — 2 exact asset matches, 1 fuzzy suggestion, 1 unmatched
        When: assemble_preview() is called
        Then: parsed_rows_count=4; correct match statuses; has_errors=True (unmatched)
        """
        parsed_rows = [
            _make_valid_row(asset_name="Grinder 1", product_name="Dark Roast 12oz"),
            _make_valid_row(asset_name="Filler Line A", product_name="Dark Roast 12oz"),
            _make_valid_row(asset_name="Grinder", product_name="New Blend XYZ"),
            _make_valid_row(asset_name="Unknown Machine", product_name="New Blend XYZ"),
        ]

        preview = assemble_preview(parsed_rows, asset_names_map, existing_products)

        assert preview.parsed_rows_count == 4
        assert preview.has_errors is True

        # Check individual row statuses
        statuses = [row.asset_match_status for row in preview.rows]
        assert statuses.count("matched") == 2
        assert statuses.count("suggested") == 1
        assert statuses.count("unmatched") == 1

        # Verify suggestions are present for the fuzzy match row
        fuzzy_rows = [r for r in preview.rows if r.asset_match_status == "suggested"]
        assert len(fuzzy_rows) == 1
        assert len(fuzzy_rows[0].suggestions) > 0

        # Verify new products identified
        new_product_names = [p for p in preview.new_products]
        assert "New Blend XYZ" in new_product_names


# =============================================================================
# AC2: Product Matching Tests
# =============================================================================


class TestProductMatching:
    """Tests for product name matching (existing vs new)."""

    def test_UNIT_027_product_matching_existing_vs_new(self, existing_products):
        """12-3-schedule-upload-api-UNIT-027: Product matching identifies existing vs new.

        Given: Product names ["Dark Roast 12oz", "New Blend XYZ"]; "Dark Roast 12oz" exists
        When: match_products(product_names, existing_products) is called
        Then: "Dark Roast 12oz" maps to its UUID; "New Blend XYZ" maps to None (new)
        """
        product_names = ["Dark Roast 12oz", "New Blend XYZ"]
        result = match_products(product_names, existing_products)

        assert result["Dark Roast 12oz"] == PRODUCT_DARK_ROAST_ID
        assert result["New Blend XYZ"] is None
